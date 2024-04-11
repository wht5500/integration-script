# -*- conding:utf-8 -*-
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


def genarate_xlsx(projectname,branch,startdate,enddate,addline, delline):
    # 创建新的Workbook对象
    wb = Workbook()

    # 选择活动的工作表或创建一个新的工作表
    ws = wb.active
    ws.title = "合计"

    # 设置默认单元格格式
    default_font = Font(name='Times New Roman', size=11)
    default_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # 设置默认边框样式
    thin_border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))

    # 设置默认对齐样式
    alignment = Alignment(horizontal='center', vertical='center')

    # 设置统一的行高，
    default_row_height = 12.25
    for row in ws.iter_rows():
        row_idx = row[0].row
        ws.row_dimensions[row_idx].height = default_row_height

    # 设置统一的列宽
    default_column_width = 14.25
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = default_column_width


    # 设置B9到J13的边框
    for row in ws.iter_rows(min_row=9, max_row=13, min_col=2, max_col=10):
        for cell in row:
            cell.border = thin_border

    # 填充数据和设置特定单元格格式
    gray_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    bold_font = Font(bold=True)

    # 项目名称
    ws['B9'] = '项目名称'
    ws['B9'].fill = gray_fill
    ws['B9'].font = bold_font
    ws['B9'].border = thin_border
    ws['B9'].alignment = alignment

    # 长安跨越一体机
    ws.merge_cells('C9:J9')
    ws['C9'] = projectname
    ws['C9'].fill = default_fill
    ws['C9'].font = default_font
    ws['C9'].border = thin_border
    ws['C9'].alignment = alignment
    ws['C9'].alignment = Alignment(horizontal='left')

    # 时间
    ws['B10'] = '时间'
    ws['B10'].fill = gray_fill
    ws['B10'].font = bold_font
    ws['B10'].border = thin_border
    ws['B10'].alignment = alignment
    ws.merge_cells('C10:J10')
    ws['C10'] = startdate + " ~ " + enddate
    ws['C10'].fill = default_fill
    ws['C10'].font = default_font
    ws['C10'].border = thin_border
    ws['C10'].alignment = Alignment(horizontal='left')



    # 类别, 分支, 增加行, 删减行, 说明
    ws.merge_cells('B11:C11')
    ws['B11'] = '类别'
    ws['B11'].fill = gray_fill
    ws['B11'].font = bold_font
    ws['B11'].border = thin_border
    ws['B11'].alignment = alignment

    ws['D11'] = '分支'
    ws['D11'].fill = gray_fill
    ws['D11'].font = bold_font
    ws['D11'].border = thin_border
    ws['D11'].alignment = alignment

    ws['E11'] = '增加行'
    ws['E11'].fill = gray_fill
    ws['E11'].font = bold_font
    ws['E11'].border = thin_border
    ws['E11'].alignment = alignment

    ws['F11'] = '删减行'
    ws['F11'].fill = gray_fill
    ws['F11'].font = bold_font
    ws['F11'].border = thin_border
    ws['F11'].alignment = alignment

    ws.merge_cells('G11:J11')
    ws['G11'] = '说明'
    ws['G11'].fill = gray_fill
    ws['G11'].font = bold_font
    ws['G11'].border = thin_border
    ws['G11'].alignment = alignment


    # MCU, develop, 增加行和删减行的数据
    ws.merge_cells('B12:C12')
    ws['B12'] = 'MCU'
    ws['B12'].border = thin_border
    ws['B12'].alignment = alignment

    # develop数据
    ws['D12'] = branch
    ws['D12'].border = thin_border
    ws['D12'].alignment = alignment

    # 设置增加行和删减行的单元格灰色填充
    gray_cell_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

    # 增加行数据
    ws['E12'] = addline
    ws['E12'].border = thin_border
    ws['E12'].alignment = alignment

    # 删减行数据
    ws['F12'] = delline
    ws['F12'].border = thin_border
    ws['F12'].alignment = alignment

    ws.merge_cells('G12:J12')
    ws['G12'].border = thin_border
    ws['G12'].alignment = alignment

    # 合计行
    ws.merge_cells('B13:D13')
    ws['B13'] = '合计'
    ws['B13'].border = thin_border
    ws['B13'].alignment = alignment

    # 增加行数据
    ws['E13'] = addline
    ws['E13'].border = thin_border
    ws['E13'].alignment = alignment

    # 删减行数据
    ws['F13'] = delline
    ws['F13'].border = thin_border
    ws['F13'].alignment = alignment


    # 合并G13到J13的单元格
    ws.merge_cells('G13:J13')
    ws['G13'] = '---'
    ws['G13'].border = thin_border
    ws['G13'].alignment = alignment

    # 保存文件
    wb.save('{}({}~{}).xlsx'.format(projectname,startdate,enddate))

    print('Excel文件已创建并保存。')

def send_mail():
    # 设置SMTP服务器地址和端口
    smtp_server = 'IP'
    port = 25  # 使用TLS

    # 设置发件人和收件人的电子邮件地址
    sender_email = ""
    receiver_email = ""
    password = ""  # 或者使用应用专用密码

    # 创建邮件对象
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receiver_email
    message['Subject'] = '[长安跨越一体机] 项目代码修改量数据统计'

    # 邮件正文内容



    body = '王工，\n\n\t附件是长安跨越一体机项目在 2024-02-01~2024-03-01期间的变更代码统计量--\n签名:你的名字'
    message.attach(MIMEText(body, 'plain'))

    # 添加附件
    #with open('长安跨越一体机(2024-02-01~2024-04-01).xlsx', 'rb') as attachment:
    #    part = MIMEBase('application', 'octet-stream')
    #    part.set_payload(attachment.read())
    #    encoders.encode_base64(part)
    #    part.add_header('Content-Disposition', 'attachment; filename= "attachment.txt"')
    #    message.attach(part)

    # 创建SMTP对象并连接到服务器
    print("test")
    with smtplib.SMTP(smtp_server, port) as server:
        server.starttls()  # 启动TLS模式
        server.login(sender_email, password)  # 登录到你的邮箱账户
        text = message.as_string()
        server.sendmail(sender_email, receiver_email, text)

    print("邮件发送成功")

def main():
    parser = argparse.ArgumentParser(description='code statistics')
    parser.add_argument('--projectname','-p',type=str,required=True,default="长安跨越一体机",help="project name")
    parser.add_argument('--branch','-b',type=str,required=True,default="master",help="branch name")
    parser.add_argument('--startdate','-s',type=str,required=True,default="2024-01-01",help="start date")
    parser.add_argument('--enddate','-e',type=str,required=True,default="2024-02-01",help="end date")
    parser.add_argument('--addline','-a',type=str,required=True,default="0",help="add line")
    parser.add_argument('--delline','-d',type=str,required=True,default="0",help="delete line")

    args = parser.parse_args()

    genarate_xlsx(args.projectname,args.branch,args.startdate,args.enddate,args.addline,args.delline)
    #send_mail()

if __name__ == "__main__":
    main()
