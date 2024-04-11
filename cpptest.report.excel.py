# -*- conding:utf-8 -*-
import argparse
import xml.etree.ElementTree as ET
import re
import json
import os,sys
import shutil
import subprocess  
import platform
import xlwt
import openpyxl
from openpyxl.styles import PatternFill  


XLS_NAME = "MISRAC2012.xlsx"
SHEET_NAME = "Info"
category_json = []

def get_rule_severity(sev):
    if '1' == sev:
        return 'Highest'
    elif '2' == sev:
        return 'High'
    elif '3' == sev:
        return 'Medium'
    else:
        return 'Low'

def get_rule_categories(xmlfile):
    print('get categories of {}'.format(xmlfile))
    tree = ET.parse(xmlfile)
    root = tree.getroot()
    
    for CodingStandard in root.find('CodingStandards'):
        if 'Rules' == CodingStandard.tag:
            for categories in CodingStandard.find('CategoriesList'):
                #print(categories.get('name'))
                if "MISRAC2012" == categories.get('name'):
                    for category in categories.findall('Category'):
                        #print(category.get('desc'))
                        category_level = category.get('desc').split('(')[1].split(')')[0]
                        category_json.append({'name':category.get('name'),'level':category_level})
                        
    #print(category_json)                    
         

def get_rule_category_level(categoryname):
    for i in category_json:
        if categoryname == i['name']:
            #print(i['level'])
            return i['level']

def get_xml_data(filename, module, xlsfile):
    print('get xml data of {}'.format(module))
    tree = ET.parse(filename)
    root = tree.getroot()

    if not os.path.exists(xlsfile):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 80
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 5
        sheet.column_dimensions['G'].width = 25
        sheet.column_dimensions['H'].width = 10 
        sheet.column_dimensions['I'].width = 10 

        xlhead = ['No.','Rule', 'Description', 'File', 'Module', 'Line', 'Category', 'Category_Level', 'Severity']
        sheet.append(xlhead)
        linecount = len(list(sheet.rows))
        for CodingStandard in root.find('CodingStandards'):
            for StdViol in CodingStandard.findall('StdViol'):
                category_level = get_rule_category_level(StdViol.get('cat'))
                #print(linecount,StdViol.get('rule'),StdViol.get('msg'),StdViol.get('locFile'),module,StdViol.get('ln'),StdViol.get('cat'),category_level,get_rule_severity(StdViol.get('sev')))
                issue = [linecount,StdViol.get('rule'),StdViol.get('msg'),StdViol.get('locFile'),module,StdViol.get('ln'),StdViol.get('cat'),category_level,get_rule_severity(StdViol.get('sev'))]
                sheet.append(issue)
                linecount = linecount + 1
        workbook.save(xlsfile)
    else:
        workbook = openpyxl.load_workbook(xlsfile)
        sheet = workbook[SHEET_NAME]
        linecount = len(list(sheet.rows))

        for CodingStandard in root.find('CodingStandards'):
            for StdViol in CodingStandard.findall('StdViol'):
                category_level = get_rule_category_level(StdViol.get('cat'))
                #print(linecount,StdViol.get('rule'),StdViol.get('msg'),StdViol.get('locFile'),module,StdViol.get('ln'),StdViol.get('cat'),category_level,get_rule_severity(StdViol.get('sev')))
                issue = [linecount,StdViol.get('rule'),StdViol.get('msg'),StdViol.get('locFile'),module,StdViol.get('ln'),StdViol.get('cat'),category_level,get_rule_severity(StdViol.get('sev'))]
                sheet.append(issue)
                linecount = linecount + 1
                
        workbook.save(xlsfile)
    


def main():
    parser = argparse.ArgumentParser(description='argparse xml parament')
    parser.add_argument('--dirname','-d',type=str,required=False,default="default.xml",help="xml file name")
    args = parser.parse_args()

    file_path = os.path.abspath(args.dirname)

    xls_file = os.path.join(file_path,XLS_NAME)
    if os.path.exists(xls_file):
        os.remove(xls_file)

    for files in os.listdir(file_path):
        if files.endswith(".xml"):
            xml = os.path.join(file_path, files)
            category_json.clear()
            get_rule_categories(xml)
            get_xml_data(xml,files.replace('_report.xml',''),xls_file) 
    print('finished!')
            


if __name__ == "__main__":
    main()

