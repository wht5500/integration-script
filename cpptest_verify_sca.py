# -*- conding:utf-8 -*-
import argparse
import subprocess
import xml.etree.ElementTree as ET
import os,sys
import xlwt
import openpyxl

CURRENT_PATH = os.getcwd()
ROOT_PATH = os.path.join(CURRENT_PATH, "..\..\..")
GERRIT_HOST = "IP"
GERRIT_PORT = "29418"

Msg_Index = 3
File_Index = 4
Line_Index = 6
Category_Index = 7
Level_Index = 8
SCA_Flag = 0



change_list = []
XLS_NAME = "MISRAC2012.xlsx"

def replace_char(msg):
    msg = msg.replace(" ", "\ ")
    msg = msg.replace("'", "")
    msg = msg.replace(' || ', '||')
    msg = msg.replace(' &', '&')
    msg = msg.replace('& ', '&')
    msg = msg.replace(' *', '*')
    msg = msg.replace('* ', '*')
    return msg

def get_column_index(column_name,xlsxpath):
    workbook = openpyxl.load_workbook(xlsxpath)  
    worksheet = workbook.active  
        
    #for col_num, column in enumerate(worksheet.iter_cols(), start=1):  
    #    for cell in column:  
    #        if cell.value == column_name:  
    #            column_index = col_num  
    #            break  
    column_max = 10
    for i in range(1,column_max): 
        print(worksheet.cell(1, i).value)
        if column_name == worksheet.cell(1, i).value:
            column_index = i
            print(i)
            return column_index

def send_gerrit_final_msg(commitid):
    global SCA_Flag

    if 0 == SCA_Flag:
        cmd = 'ssh -p {} {} gerrit review --label Sonar-Verified=+1 -m "looks\ good\ for\ me!" {}'.format(GERRIT_PORT,GERRIT_HOST,commitid)
    elif 1 == SCA_Flag:
        cmd = 'ssh -p {} {} gerrit review --label Sonar-Verified=+1 -m "find\ some\ medium\ or\ low\ level\ issue,\ please\ check\ above,\ and\ decide\ whether\ it\ is\ necessary\ to\ fix!" {}'.format(GERRIT_PORT,GERRIT_HOST,commitid)
    elif 2 == SCA_Flag:
        cmd = 'ssh -p {} {} gerrit review --label Sonar-Verified=-1 -m "find\ some\ high\ level\ issue,\ please\ check\ above,\ and\ fix\ it!!!" {}'.format(GERRIT_PORT,GERRIT_HOST,commitid)
    print(cmd)
    retcode,output = subprocess.getstatusoutput(cmd)
    if 0 != retcode:
        print(output)
        exit(1)


def send_gerrit_verify_msg(commitid,issuemsg,level,line,issuecategory,filename):
    global SCA_Flag
    
    if "Mandatory" == level:
        cmd = 'ssh -p {} {} gerrit review --label Sonar-Verified=-1 -m "FILE:{},\ LINE:{},\ RULE:{},\ CATETOGORY:{},\ MSG:{}" {}'.format(GERRIT_PORT,GERRIT_HOST,filename,line,issuecategory,level,replace_char(issuemsg),commitid)
        print(cmd)
        SCA_Flag = 2 #find high level issues
        retcode,output = subprocess.getstatusoutput(cmd)
        if 0 != retcode:
            print(output)
            exit(1)
    else:
        cmd = 'ssh -p {} {} gerrit review -m "FILE:{},\ LINE:{},\ RULE:{},\ CATETOGORY:{},\ MSG:{}" {}'.format(GERRIT_PORT,GERRIT_HOST,filename,line,issuecategory,level,replace_char(issuemsg),commitid)
        print(cmd)
        if SCA_Flag < 2:
            SCA_Flag = 1 #find medium and low level issues
        retcode,output = subprocess.getstatusoutput(cmd)
        if 0 != retcode:
            print(output)
            exit(1)
    
def check_issue(filename,start_index,length,commitid,xlsxpath):
    
    if not os.path.exists(xlsxpath):
        print('code statics analysis result not exists!')
        return
    else:
        workbook = openpyxl.load_workbook(xlsxpath)  
        worksheet = workbook.active  
        rowindex = 2

        for row in worksheet.iter_rows():  
            cell_value = worksheet.cell(rowindex, File_Index).value
            if str(filename).split('/')[-1] == str(cell_value).split('/')[-1]:
                issueline = worksheet.cell(rowindex, Line_Index).value

                if int(start_index) <= int(issueline) and int(issueline) <= (int(start_index) + int(length)):
                    issuelevel = worksheet.cell(rowindex, Level_Index).value
                    issuecategory= worksheet.cell(rowindex, Category_Index).value
                    issuemsg= worksheet.cell(rowindex, Msg_Index).value
                    print(cell_value,issueline,issuecategory,issuelevel,issuemsg)
                    send_gerrit_verify_msg(commitid,issuemsg,issuelevel,issueline,issuecategory,filename)
            rowindex = rowindex + 1
                

    

    
def get_changefile():
    cmd = 'git diff --name-status HEAD~1 HEAD'
    retcode,output = subprocess.getstatusoutput(cmd)
    for i in output.split('\n'):
        changemode = i[0]
        changefile = i.split('\t')[1]
        change_list.append({'name':changefile,'mode':changemode})

def get_changesection(commitid,xlsxpath):
    for i in change_list:
        if 'D' != i['mode'] and  i['name'].endswith(('.c','.h')):
            cmd = 'git diff -U0 HEAD~1 HEAD "../../../{}" |findstr /r "^@@ -(\d+),(\d+) \+(\d+),(\d+) @@$"'.format(i['name'])
            retcode,output = subprocess.getstatusoutput(cmd)
            for sindex in output.split('\n'):
                startstr = sindex.find('+')  
                endstr = sindex.find(' @@')
                rangenum = sindex[startstr + 1:endstr]
                if ',' in rangenum:
                    start_index = rangenum.split(',')[0]
                    length = rangenum.split(',')[1]
                else:
                    start_index = rangenum
                    length = 1
                print("checking file:",i['name'],"range:",start_index,length)
                check_issue(i['name'],start_index,length,commitid,xlsxpath)
    



def main():
    parser = argparse.ArgumentParser(description='argparse xml parament')
    parser.add_argument('--commitid','-i',type=str,required=True,default="",help="commit id")
    parser.add_argument('--path','-p',type=str,required=True,default="",help="xlsx path")
    args = parser.parse_args()

    commitid = args.commitid
    xlsxpath = os.path.join(args.path,XLS_NAME)

    get_changefile()
    get_changesection(commitid,xlsxpath)
    send_gerrit_final_msg(commitid)


if __name__ == "__main__":
    main()
